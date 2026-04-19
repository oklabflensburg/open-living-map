import logging
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text
from sqlmodel import select

from app.core.ars import normalize_ars
from app.core.config import settings
from app.etl.common import clear_indicator_values, get_or_create_indicator, normalize, with_session
from app.models.indicator import RegionIndicatorValue
from app.models.region import Region

logger = logging.getLogger("etl.import_flaechenatlas")
logging.basicConfig(level=logging.INFO)

XLSX_PATH = Path(__file__).resolve().parents[3] / "data" / "raw" / "flaechenatlas2019daten.xlsx"


def _parse_float(value: object) -> float | None:
    stripped = str(value).strip()
    if not stripped:
        return None
    try:
        return float(stripped.replace(",", "."))
    except ValueError:
        return None


def ensure_flaechenatlas_table(session) -> None:
    session.execute(text("CREATE SCHEMA IF NOT EXISTS landuse"))
    session.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS landuse.region_area_stat (
                region_ars text PRIMARY KEY,
                year integer NOT NULL,
                forest_share_pct double precision,
                settlement_transport_share_pct double precision,
                agriculture_share_pct double precision,
                transport_share_pct double precision,
                settlement_transport_sqm_per_capita double precision
            )
            """
        )
    )
    session.execute(
        text(
            """
            CREATE INDEX IF NOT EXISTS region_area_stat_year_idx
            ON landuse.region_area_stat (year)
            """
        )
    )
    session.commit()


def _batch_write_indicator_values(
    session,
    *,
    indicator_id: int,
    period: str,
    values: list[tuple[int, float]],
    normalized_values: list[float],
) -> None:
    existing_rows = {
        row.region_id: row
        for row in session.exec(
            select(RegionIndicatorValue).where(
                RegionIndicatorValue.indicator_id == indicator_id,
                RegionIndicatorValue.period == period,
            )
        )
    }

    for (region_id, raw_value), normalized_value in zip(values, normalized_values):
        existing = existing_rows.get(region_id)
        if existing:
            existing.raw_value = round(raw_value, 4)
            existing.normalized_value = normalized_value
            existing.quality_flag = "ok"
            session.add(existing)
            continue

        session.add(
            RegionIndicatorValue(
                region_id=region_id,
                indicator_id=indicator_id,
                period=period,
                raw_value=round(raw_value, 4),
                normalized_value=normalized_value,
                quality_flag="ok",
            )
        )


def parse_flaechenatlas_rows(path: Path) -> list[dict[str, float | int | str | None]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows: list[dict[str, float | int | str | None]] = []
    for values in sheet.iter_rows(min_row=4, max_col=7, values_only=True):
        ars = normalize_ars(str(values[0] or ""))
        if len(ars) != 8:
            continue

        rows.append(
            {
                "region_ars": ars,
                "year": 2019,
                "forest_share_pct": _parse_float(values[2] or ""),
                "settlement_transport_share_pct": _parse_float(values[3] or ""),
                "agriculture_share_pct": _parse_float(values[4] or ""),
                "transport_share_pct": _parse_float(values[5] or ""),
                "settlement_transport_sqm_per_capita": _parse_float(values[6] or ""),
            }
        )
    workbook.close()
    return rows


def main() -> None:
    logger.info("Flächenatlas-Import gestartet")
    if not XLSX_PATH.exists():
        logger.warning("Flächenatlas-Datei nicht gefunden: %s", XLSX_PATH)
        return

    rows = parse_flaechenatlas_rows(XLSX_PATH)
    if not rows:
        logger.warning("Flächenatlas-Datei lieferte keine parsebaren Gemeindedaten.")
        return

    with with_session() as session:
        ensure_flaechenatlas_table(session)
        session.execute(text("TRUNCATE landuse.region_area_stat"))
        session.execute(
            text(
                """
                INSERT INTO landuse.region_area_stat (
                    region_ars,
                    year,
                    forest_share_pct,
                    settlement_transport_share_pct,
                    agriculture_share_pct,
                    transport_share_pct,
                    settlement_transport_sqm_per_capita
                )
                VALUES (
                    :region_ars,
                    :year,
                    :forest_share_pct,
                    :settlement_transport_share_pct,
                    :agriculture_share_pct,
                    :transport_share_pct,
                    :settlement_transport_sqm_per_capita
                )
                """
            ),
            rows,
        )
        session.commit()

        regions_by_ars = {
            region.ars: region.id
            for region in session.exec(select(Region).where(Region.level == "gemeinde"))
        }

        indicator_specs = [
            (
                "forest_share_pct",
                "forest_share_pct",
                "Waldanteil",
                "higher_is_better",
            ),
            (
                "settlement_transport_share_pct",
                "settlement_transport_share_pct",
                "Siedlungs- und Verkehrsflächenanteil",
                "lower_is_better",
            ),
            (
                "agriculture_share_pct",
                "agriculture_share_pct",
                "Landwirtschaftsanteil",
                "higher_is_better",
            ),
            (
                "transport_share_pct",
                "transport_share_pct",
                "Verkehrsflächenanteil",
                "lower_is_better",
            ),
            (
                "settlement_transport_sqm_per_capita",
                "settlement_transport_sqm_per_capita",
                "Siedlungs- und Verkehrsfläche je Einwohner",
                "lower_is_better",
            ),
        ]

        source_url = "https://www.destatis.de/DE/Service/Statistik-Visualisiert/flaechenatlas.html"
        methodology = (
            "Flächenatlas 2019 auf Gemeindeebene: Waldanteil, Landwirtschaftsanteil, "
            "Siedlungs- und Verkehrsflächenanteil, Verkehrsflächenanteil sowie "
            "Siedlungs- und Verkehrsfläche je Einwohner."
        )

        for field_name, indicator_key, indicator_name, direction in indicator_specs:
            indicator = get_or_create_indicator(
                session,
                key=indicator_key,
                name=indicator_name,
                category="landuse",
                unit="sqm_per_capita" if field_name == "settlement_transport_sqm_per_capita" else "percent",
                direction=direction,
                normalization_mode="robust_percentile",
                source_name="Destatis Flächenatlas 2019",
                source_url=source_url,
                methodology=methodology,
            )
            clear_indicator_values(
                session,
                indicator_id=indicator.id,
                period=settings.default_score_period,
            )

            values = [
                (regions_by_ars[row["region_ars"]], float(row[field_name]))
                for row in rows
                if row["region_ars"] in regions_by_ars and row[field_name] is not None
            ]
            if not values:
                continue

            normalized_values = normalize(
                [raw_value for _, raw_value in values],
                direction,
                mode="robust_percentile",
            )
            _batch_write_indicator_values(
                session,
                indicator_id=indicator.id,
                period=settings.default_score_period,
                values=values,
                normalized_values=normalized_values,
            )

        session.commit()

    logger.info("Flächenatlas-Import abgeschlossen (%s Gemeinden)", len(rows))


if __name__ == "__main__":
    main()
